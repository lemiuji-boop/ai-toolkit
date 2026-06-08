# Copyright 2026 Rinat Ishmaev
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string


def _col_letter(mapping: dict, key: str, default: int) -> int:
    """Колонка из mapping (буква A,B) или номер по умолчанию."""
    val = (mapping.get("columns") or {}).get(key)
    if isinstance(val, str) and val.isalpha():
        return column_index_from_string(val.upper())
    return default


def export_materials_to_excel(
    template_bytes: bytes | None,
    rows: list[dict],
    field_mapping: dict | None = None,
) -> bytes:
    """Экспорт ведомости материалов с mapping из EXCEL_TEMPLATE_SPEC."""
    mapping = field_mapping or {}
    sheet_name = mapping.get("sheet")
    start_row = int(mapping.get("start_row", 5))

    if template_bytes:
        wb = load_workbook(BytesIO(template_bytes))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        cols = mapping.get("columns") or {}
        c_des = _col_letter(mapping, "designation", 1)
        c_mat = _col_letter(mapping, "material", 2)
        c_net = _col_letter(mapping, "net_qty", 3)
        c_gross = _col_letter(mapping, "gross_qty", 4)
        c_unit = _col_letter(mapping, "unit", 5)
        c_kim = _col_letter(mapping, "kim", 6) if "kim" in cols else None
        for i, row in enumerate(rows):
            r = start_row + i
            ws.cell(row=r, column=c_des, value=row.get("designation", ""))
            ws.cell(row=r, column=c_mat, value=row.get("material_name", row.get("material", "")))
            ws.cell(row=r, column=c_net, value=row.get("net_qty"))
            ws.cell(row=r, column=c_gross, value=row.get("gross_qty"))
            ws.cell(row=r, column=c_unit, value=row.get("unit", "kg"))
            if c_kim:
                ws.cell(row=r, column=c_kim, value=row.get("kim"))
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name or "Материалы"
        headers = ["Обозначение", "Материал", "Нетто", "Брутто", "Ед.", "КИМ"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for i, row in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=row.get("designation", ""))
            ws.cell(row=i, column=2, value=row.get("material_name", ""))
            ws.cell(row=i, column=3, value=row.get("net_qty"))
            ws.cell(row=i, column=4, value=row.get("gross_qty"))
            ws.cell(row=i, column=5, value=row.get("unit", "kg"))
            ws.cell(row=i, column=6, value=row.get("kim"))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
