"""Серверная выгрузка ведомости материалов в .xlsx (опциональный путь;
основной перенос в Excel делает надстройка через Office.js)."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.schemas import NormRow

HEADERS = [
    "№", "Обозначение", "Наименование", "Материал", "Вид заготовки",
    "Кол-во на с/к", "Мд, кг", "Мз, кг", "КИМ",
    "Норма на деталь, кг", "Норма на программу, кг", "Замечания",
]


def to_xlsx(rows: list[NormRow]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость материалов"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
    warn = PatternFill("solid", fgColor="FAEEDA")
    for r in rows:
        ws.append([
            r.num, r.designation, r.name, r.material, r.zagotovka,
            r.qty_per_set, r.md_kg, r.mz_kg, r.kim,
            r.norm_per_part_kg, r.norm_program_kg, "; ".join(r.flags),
        ])
        if r.flags:
            ws.cell(row=ws.max_row, column=len(HEADERS)).fill = warn
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
