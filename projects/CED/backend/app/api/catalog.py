# Copyright 2026 Rinat Ishmaev
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from datetime import date
from io import BytesIO
import re

from fastapi import APIRouter, Depends, Query
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.config import settings
from app.core.database import get_db_session
from app.models.change_order import ChangeOrder
from app.models.change_order_directive import ChangeOrderDirective
from app.models.document_version import DocumentVersion
from app.models.user import User
from app.schemas.catalog import AiStatusOut, CatalogListResponse, CatalogRowOut
from app.services.analysis_service import analysis_to_ui_status, get_latest_analysis
from app.services.catalog_service import build_catalog_query, paginate_catalog

router = APIRouter()


async def _directives_for_document(
    db: AsyncSession, document_id: int, doc_number: str
) -> list[tuple[ChangeOrder, ChangeOrderDirective]]:
    """Связанные ИИ и директивы для строки каталога."""
    stmt = (
        select(ChangeOrder, ChangeOrderDirective)
        .join(ChangeOrderDirective, ChangeOrderDirective.change_order_id == ChangeOrder.id)
        .where(ChangeOrder.deleted_at.is_(None))
        .where(
            (ChangeOrderDirective.document_id == document_id)
            | (ChangeOrderDirective.affected_doc_number == doc_number)
        )
    )
    return list((await db.execute(stmt)).all())


@router.get("/catalog", response_model=CatalogListResponse)
async def list_catalog(
    doc_number: str | None = None,
    name: str | None = None,
    order_number: str | None = None,
    backlog: str | None = None,
    impl_date_from: date | None = None,
    impl_date_to: date | None = None,
    product: str | None = None,
    status: str | None = None,
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=500),
    db: AsyncSession = Depends(get_db_session),
    _: User = Depends(get_current_user),
) -> CatalogListResponse:
    stmt = build_catalog_query(
        doc_number=doc_number,
        name=name,
        order_number=order_number,
        backlog=backlog,
        impl_date_from=impl_date_from,
        impl_date_to=impl_date_to,
        product=product,
        status=status,
    )
    rows, total = await paginate_catalog(db, stmt, page, size)
    items: list[CatalogRowOut] = []
    for row in rows:
        version_index = None
        analysis = await get_latest_analysis(db, row.current_version_id)
        if row.current_version_id:
            version = await db.get(DocumentVersion, row.current_version_id)
            if version:
                version_index = version.version_index
        ui = analysis_to_ui_status(
            analysis,
            has_version=bool(row.current_version_id),
            ocr_processed=row.ocr_processed,
        )
        insight = None
        if analysis and analysis.extracted_data:
            insight = analysis.extracted_data.get("insight") or analysis.error_message
        items.append(
            CatalogRowOut(
                id=row.id,
                doc_number=row.doc_number,
                name=row.name,
                product_number=row.product_number,
                doc_type=row.doc_type,
                status=row.status.value if hasattr(row.status, "value") else str(row.status),
                catalog_path=row.catalog_path,
                version_index=version_index,
                updated_at=row.updated_at,
                ai_status=AiStatusOut(**ui),
                ai_insight=insight,
            )
        )
    return CatalogListResponse(items=items, total_count=total)


@router.get("/catalog/export")
async def export_catalog(
    doc_number: str | None = None,
    name: str | None = None,
    order_number: str | None = None,
    backlog: str | None = None,
    impl_date_from: date | None = None,
    impl_date_to: date | None = None,
    product: str | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db_session),
    _: User = Depends(get_current_user),
) -> StreamingResponse:
    stmt = build_catalog_query(
        doc_number=doc_number,
        name=name,
        order_number=order_number,
        backlog=backlog,
        impl_date_from=impl_date_from,
        impl_date_to=impl_date_to,
        product=product,
        status=status,
    )
    rows, _ = await paginate_catalog(db, page=1, size=10000, stmt=stmt)

    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"
    ws.append(
        [
            "Обозначение",
            "Наименование",
            "Изделие",
            "Версия",
            "№ ИИ",
            "Задел",
            "Срок/изделие",
            "Статус",
            "Путь",
            "Тип строки",
        ]
    )

    current_group: str | None = None
    group_start = 2
    for row in rows:
        group_key = row.catalog_path or row.product_number or "прочее"
        if current_group != group_key:
            if current_group is not None:
                ws.row_dimensions.group(group_start, ws.max_row, outline_level=1)
            current_group = group_key
            group_start = ws.max_row + 1

        version_index = ""
        if row.current_version_id:
            version = await db.get(DocumentVersion, row.current_version_id)
            if version:
                version_index = version.version_index

        ws.append(
            [
                row.doc_number,
                row.name,
                row.product_number,
                version_index,
                "",
                "",
                "",
                str(row.status),
                row.catalog_path,
                "КД",
            ]
        )

        links = await _directives_for_document(db, row.id, row.doc_number)
        for co, directive in links:
            impl = ""
            if directive.implementation_kind:
                impl = directive.implementation_kind.value
            if directive.implementation_product:
                impl = f"{impl} / {directive.implementation_product}".strip(" /")
            if directive.implementation_date:
                impl = f"{impl} / {directive.implementation_date}".strip(" /")
            ws.append(
                [
                    directive.affected_doc_number or row.doc_number,
                    directive.directive_text or co.description,
                    directive.implementation_product,
                    "",
                    co.order_number,
                    directive.backlog_action.value if directive.backlog_action else "",
                    impl,
                    "ИИ",
                    co.file_path,
                    "директива",
                ]
            )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalog_export.xlsx"},
    )


class ChatRequest(BaseModel):
    message: str


@router.post("/catalog/chat")
async def catalog_chat(
    payload: ChatRequest,
    db: AsyncSession = Depends(get_db_session),
    _: User = Depends(get_current_user),
) -> dict:
    text = payload.message.strip()
    low = text.lower()
    filters: dict = {}
    actions: list[dict] = []

    doc_match = re.search(r"([а-яА-Яa-zA-Z0-9]+[.\-][\w.\-]+)", text)
    if doc_match:
        filters["doc_number"] = doc_match.group(1)
    if "наименован" in low:
        words = [w for w in text.split() if len(w) > 3][-3:]
        if words:
            filters["name"] = " ".join(words)
    if "ии" in low or "извещ" in low:
        num = re.search(r"ии[-\s]?(\d+)", low)
        if num:
            filters["order_number"] = f"ИИ-{num.group(1)}"

    stmt = build_catalog_query(
        doc_number=filters.get("doc_number"),
        name=filters.get("name"),
        order_number=filters.get("order_number"),
    )
    rows, total = await paginate_catalog(db, stmt, 1, 20)
    items = [
        {"id": r.id, "doc_number": r.doc_number, "name": r.name, "catalog_path": r.catalog_path}
        for r in rows
    ]

    if ("путь" in low or "папк" in low or "откр" in low) and rows:
        doc = rows[0]
        full = f"{settings.catalog_root}/{doc.catalog_path}" if doc.catalog_path else str(settings.catalog_root)
        actions = [
            {"type": "open_folder", "path": full, "document_id": doc.id},
            {"type": "open_file", "path": full, "document_id": doc.id},
        ]
        return {
            "reply": f"Каталог КД {doc.doc_number}: {full}",
            "filters": filters or None,
            "items": items,
            "actions": actions,
        }

    if not items:
        return {"reply": "По запросу ничего не найдено.", "filters": filters or None, "items": [], "actions": []}

    return {
        "reply": f"Найдено: {total}. Уточните обозначение для открытия папки.",
        "filters": filters or None,
        "items": items,
        "actions": actions,
    }
